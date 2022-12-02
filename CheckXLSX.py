import openpyxl
def CheckXLSX(url):
    wb_obj = openpyxl.load_workbook("output/output.xlsx")
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    ls = []
    for i in range(2, m_row + 1):
        ls.append(sheet_obj.cell(row=i, column=11).value)
        # print(sheet_obj.cell(row=i, column=11).value)
    if url in ls:
        return True
    else:
        return False
