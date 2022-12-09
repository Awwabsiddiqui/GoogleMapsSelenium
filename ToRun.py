import openpyxl

from main import mainMethod

wb_obj = openpyxl.load_workbook("input/input.xlsx")
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row

for i in range(2, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    size = sheet_obj.cell(row=i, column=2)
    category = sheet_obj.cell(row=i, column=3)
    print(cell_obj.value + "  :  "+str(int(size.value)) +"  :  "+ str(category.value))
    mainMethod(cell_obj.value, 3, ((int(size.value) * 2) + 2) , category.value)

